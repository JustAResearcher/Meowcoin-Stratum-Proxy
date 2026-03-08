#!/usr/bin/env python3
"""
Meowcoin MeowPoW Solo Mining Stratum Proxy
===========================================

Bridge between ProgPow-capable GPU miners (WildRig Multi, T-Rex, GMiner, etc.)
and a Meowcoin full node via JSON-RPC.

The proxy:
  1. Polls the node's getblocktemplate RPC for new work
  2. Constructs a valid coinbase transaction (with community fund output)
  3. Computes the MeowPoW header hash and epoch seed hash
  4. Distributes work to connected miners via Stratum v1 (kawpow variant)
  5. On solution, assembles the full block and submits via submitblock

Requirements:
    pip install pycryptodome requests

Usage:
    python stratum_proxy.py --address <YOUR_MEWC_ADDRESS> [options]

    Options:
        --address       Meowcoin payout address (required)
        --rpc-host      Node RPC host           (default: 127.0.0.1)
        --rpc-port      Node RPC port           (default: 8332)
        --rpc-user      Node RPC username        (default: cookie auth)
        --rpc-pass      Node RPC password        (default: cookie auth)
        --cookie-dir    Data dir with .cookie    (default: auto-detect)
        --stratum-host  Stratum listen host      (default: 0.0.0.0)
        --stratum-port  Stratum listen port      (default: 3333)
        --poll-interval GBT poll interval (s)    (default: 1.0)
        --log-level     Logging level            (default: INFO)
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import json
import logging
import os
import struct
import sys
import time
import traceback
import uuid
from typing import Any, Dict, List, Optional, Tuple

import requests

try:
    import http.client
    import ssl
except ImportError:
    http = None
    ssl = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

from datetime import datetime, timezone

try:
    from Crypto.Hash import keccak as _keccak_module
except ImportError:
    _keccak_module = None

# =============================================================================
# Constants — Meowcoin consensus
# =============================================================================

EPOCH_LENGTH              = 7500
COMMUNITY_FUND_PCT        = 40
COMMUNITY_FUND_ADDRESS    = "MPyNGZSSZ4rbjkVJRLn3v64pMcktpEYJnU"
SUBSIDY_HALVING_INTERVAL  = 2_100_000
INITIAL_SUBSIDY_COINS     = 5000
COIN                      = 100_000_000

# Meowcoin mainnet base58 address version bytes
PUBKEY_ADDRESS_VERSION    = 50   # 'M' prefix
SCRIPT_ADDRESS_VERSION    = 122  # 'm' prefix
BECH32_HRP                = "mewc"
BECH32_HRP_REGTEST        = "rmewc"
BECH32_HRP_TESTNET        = "tmewc"
BECH32_HRPS               = (BECH32_HRP, BECH32_HRP_REGTEST, BECH32_HRP_TESTNET)

log = logging.getLogger("stratum-proxy")

# =============================================================================
# Utility — SHA-256d, Keccak-256
# =============================================================================

def sha256d(data: bytes) -> bytes:
    """Double-SHA-256 hash."""
    return hashlib.sha256(hashlib.sha256(data).digest()).digest()


def keccak256(data: bytes) -> bytes:
    """Keccak-256 hash (NOT NIST SHA3-256)."""
    if _keccak_module is not None:
        k = _keccak_module.new(digest_bits=256)
        k.update(data)
        return k.digest()
    # Fallback: try pysha3
    try:
        import sha3  # type: ignore
        return sha3.keccak_256(data).digest()
    except ImportError:
        raise RuntimeError(
            "No Keccak-256 implementation found.  "
            "Install pycryptodome:  pip install pycryptodome"
        )

# =============================================================================
# Utility — Byte / hex helpers
# =============================================================================

def hex_to_bytes(h: str) -> bytes:
    return bytes.fromhex(h)


def bytes_to_hex(b: bytes) -> str:
    return b.hex()


def reverse_hex(h: str) -> str:
    """Reverse the byte order of a hex string."""
    return bytes.fromhex(h)[::-1].hex()


def int32_le(v: int) -> bytes:
    return struct.pack("<i", v)


def uint32_le(v: int) -> bytes:
    return struct.pack("<I", v)


def uint64_le(v: int) -> bytes:
    return struct.pack("<Q", v)


def int64_le(v: int) -> bytes:
    return struct.pack("<q", v)


def compact_size(n: int) -> bytes:
    """Bitcoin varint (CompactSize) encoding."""
    if n < 0xFD:
        return struct.pack("<B", n)
    elif n <= 0xFFFF:
        return b"\xfd" + struct.pack("<H", n)
    elif n <= 0xFFFFFFFF:
        return b"\xfe" + struct.pack("<I", n)
    else:
        return b"\xff" + struct.pack("<Q", n)


def push_data(data: bytes) -> bytes:
    """Bitcoin script PUSH operation for arbitrary data."""
    n = len(data)
    if n < 0x4C:
        return bytes([n]) + data
    elif n <= 0xFF:
        return b"\x4c" + struct.pack("<B", n) + data
    elif n <= 0xFFFF:
        return b"\x4d" + struct.pack("<H", n) + data
    else:
        return b"\x4e" + struct.pack("<I", n) + data

# =============================================================================
# Utility — BIP34 height encoding
# =============================================================================

def encode_height(height: int) -> bytes:
    """
    Encode block height as a Bitcoin CScriptNum for BIP34 coinbase scriptSig.
    Returns the PUSH opcode + serialized height.
    """
    if height == 0:
        return b"\x01\x00"  # OP_PUSH1 0x00
    # Convert to signed little-endian
    neg = height < 0
    v = abs(height)
    result = []
    while v:
        result.append(v & 0xFF)
        v >>= 8
    # If high bit set, add zero byte (positive) or 0x80 (negative)
    if result[-1] & 0x80:
        result.append(0x80 if neg else 0x00)
    elif neg:
        result[-1] |= 0x80
    data = bytes(result)
    return bytes([len(data)]) + data

# =============================================================================
# Utility — Address decoding
# =============================================================================

BASE58_ALPHABET = b"123456789ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnopqrstuvwxyz"


def base58_decode(s: str) -> bytes:
    """Decode base58check string → (version_byte, payload)."""
    n = 0
    for ch in s.encode("ascii"):
        n = n * 58 + BASE58_ALPHABET.index(ch)
    # Determine required length
    byte_len = (n.bit_length() + 7) // 8
    raw = n.to_bytes(max(byte_len, 1), "big")
    # Leading '1' chars → leading zero bytes
    pad = 0
    for ch in s:
        if ch == "1":
            pad += 1
        else:
            break
    raw = b"\x00" * pad + raw
    payload, checksum = raw[:-4], raw[-4:]
    if sha256d(payload)[:4] != checksum:
        raise ValueError(f"Invalid base58check checksum for '{s}'")
    return payload  # version(1) + hash(20+)


BECH32_CHARSET = "qpzry9x8gf2tvdw0s3jn54khce6mua7l"


def _bech32_polymod(values):
    GEN = [0x3B6A57B2, 0x26508E6D, 0x1EA119FA, 0x3D4233DD, 0x2A1462B3]
    chk = 1
    for v in values:
        b = chk >> 25
        chk = ((chk & 0x1FFFFFF) << 5) ^ v
        for i in range(5):
            chk ^= GEN[i] if ((b >> i) & 1) else 0
    return chk


def _bech32_hrp_expand(hrp):
    return [ord(x) >> 5 for x in hrp] + [0] + [ord(x) & 31 for x in hrp]


def bech32_decode(bech: str):
    """Decode bech32/bech32m → (hrp, 5-bit data, spec)."""
    bech = bech.lower()
    pos = bech.rfind("1")
    if pos < 1 or pos + 7 > len(bech):
        return None, None, None
    hrp = bech[:pos]
    data = [BECH32_CHARSET.find(x) for x in bech[pos + 1 :]]
    if -1 in data:
        return None, None, None
    check = _bech32_polymod(_bech32_hrp_expand(hrp) + data)
    if check == 1:
        return hrp, data[:-6], "bech32"
    if check == 0x2BC830A3:
        return hrp, data[:-6], "bech32m"
    return None, None, None


def _convertbits(data, frombits, tobits, pad=True):
    acc, bits, ret, maxv = 0, 0, [], (1 << tobits) - 1
    for value in data:
        acc = (acc << frombits) | value
        bits += frombits
        while bits >= tobits:
            bits -= tobits
            ret.append((acc >> bits) & maxv)
    if pad:
        if bits:
            ret.append((acc << (tobits - bits)) & maxv)
    elif bits >= frombits or ((acc << (tobits - bits)) & maxv):
        return None
    return ret


def address_to_scriptpubkey(addr: str) -> bytes:
    """Convert a Meowcoin address to its scriptPubKey bytes."""
    # Bech32 / Bech32m — support mainnet, testnet, and regtest HRPs
    addr_lower = addr.lower()
    matched_hrp = None
    for candidate_hrp in BECH32_HRPS:
        if addr_lower.startswith(candidate_hrp + "1"):
            matched_hrp = candidate_hrp
            break
    if matched_hrp is not None:
        hrp, data5, spec = bech32_decode(addr)
        if hrp != matched_hrp or data5 is None:
            raise ValueError(f"Invalid bech32 address: {addr}")
        witver = data5[0]
        witprog = bytes(_convertbits(data5[1:], 5, 8, False))
        if witver == 0 and len(witprog) == 20:
            return bytes([0x00, 0x14]) + witprog      # P2WPKH
        if witver == 0 and len(witprog) == 32:
            return bytes([0x00, 0x20]) + witprog      # P2WSH
        # Witness v1+ (taproot, etc.)
        return bytes([0x50 + witver, len(witprog)]) + witprog

    # Base58Check
    raw = base58_decode(addr)
    version = raw[0]
    pubkey_hash = raw[1:]
    if version == PUBKEY_ADDRESS_VERSION and len(pubkey_hash) == 20:
        # P2PKH: OP_DUP OP_HASH160 <20> OP_EQUALVERIFY OP_CHECKSIG
        return b"\x76\xa9\x14" + pubkey_hash + b"\x88\xac"
    if version == SCRIPT_ADDRESS_VERSION and len(pubkey_hash) == 20:
        # P2SH: OP_HASH160 <20> OP_EQUAL
        return b"\xa9\x14" + pubkey_hash + b"\x87"
    raise ValueError(f"Unsupported address type (version={version}): {addr}")

# =============================================================================
# Utility — Merkle root
# =============================================================================

def merkle_root(hashes: List[bytes]) -> bytes:
    """
    Compute the Bitcoin-style merkle root from a list of 32-byte hashes
    (in internal byte order = little-endian).
    """
    if not hashes:
        return b"\x00" * 32
    layer = list(hashes)
    while len(layer) > 1:
        if len(layer) % 2 != 0:
            layer.append(layer[-1])
        next_layer = []
        for i in range(0, len(layer), 2):
            next_layer.append(sha256d(layer[i] + layer[i + 1]))
        layer = next_layer
    return layer[0]

# =============================================================================
# Utility — Epoch seed hash
# =============================================================================

_seed_cache: Dict[int, bytes] = {}


def compute_seed_hash(epoch: int) -> bytes:
    """Compute the ethash seed hash for a given epoch number."""
    if epoch in _seed_cache:
        return _seed_cache[epoch]
    seed = b"\x00" * 32
    for i in range(epoch):
        seed = keccak256(seed)
    _seed_cache[epoch] = seed
    return seed

# =============================================================================
# Utility — Block subsidy
# =============================================================================

def get_block_subsidy(height: int) -> int:
    """Returns block subsidy in satoshis for the given height."""
    halvings = height // SUBSIDY_HALVING_INTERVAL
    if halvings >= 64:
        return 0
    return (INITIAL_SUBSIDY_COINS * COIN) >> halvings

# =============================================================================
# Utility — Coinbase transaction builder
# =============================================================================

def build_coinbase_tx(
    height: int,
    miner_value: int,
    miner_script: bytes,
    community_value: int,
    community_script: bytes,
    witness_commitment: Optional[bytes] = None,
    extra_nonce: bytes = b"",
) -> bytes:
    """
    Build a serialised coinbase transaction compatible with Meowcoin consensus.

    Outputs:
        [0] miner payout
        [1] community fund (40% of subsidy)
        [2] witness commitment OP_RETURN  (if segwit)

    Returns the full serialised transaction bytes.
    """
    # --- ScriptSig: BIP34 height + optional extra ---
    script_sig = encode_height(height) + b"\x00" + extra_nonce  # OP_0 = 0x00

    n_version = 2
    n_locktime = max(height - 1, 0)
    has_witness = witness_commitment is not None

    parts: list[bytes] = []

    # nVersion
    parts.append(uint32_le(n_version))

    if has_witness:
        # Segwit marker + flag
        parts.append(b"\x00\x01")

    # --- Inputs (1 coinbase input) ---
    parts.append(compact_size(1))
    parts.append(b"\x00" * 32)                   # prev txid (null)
    parts.append(b"\xff\xff\xff\xff")             # prev vout (0xFFFFFFFF)
    parts.append(compact_size(len(script_sig)))
    parts.append(script_sig)
    parts.append(uint32_le(0xFFFFFFFE))           # nSequence (MAX_SEQUENCE_NONFINAL)

    # --- Outputs ---
    n_outputs = 2  # miner + community
    if has_witness:
        n_outputs += 1  # witness commitment OP_RETURN
    parts.append(compact_size(n_outputs))

    # Output 0: Miner payout
    parts.append(int64_le(miner_value))
    parts.append(compact_size(len(miner_script)))
    parts.append(miner_script)

    # Output 1: Community Autonomous Fund
    parts.append(int64_le(community_value))
    parts.append(compact_size(len(community_script)))
    parts.append(community_script)

    # Output 2: Witness commitment (OP_RETURN)
    if has_witness:
        parts.append(int64_le(0))
        parts.append(compact_size(len(witness_commitment)))
        parts.append(witness_commitment)

    # --- Witness data (for coinbase input) ---
    if has_witness:
        parts.append(compact_size(1))             # 1 witness field for 1 input
        parts.append(compact_size(32))            # 32-byte witness nonce
        parts.append(b"\x00" * 32)                # all-zero nonce

    # nLockTime
    parts.append(uint32_le(n_locktime))

    return b"".join(parts)

# =============================================================================
# Node RPC client
# =============================================================================

class NodeRPC:
    """JSON-RPC client for the Meowcoin node."""

    def __init__(
        self,
        host: str = "127.0.0.1",
        port: int = 8332,
        user: str = "",
        password: str = "",
        cookie_dir: str = "",
    ):
        self.url = f"http://{host}:{port}"
        self._user = user
        self._password = password
        self._cookie_dir = cookie_dir
        self._cookie_cache: Optional[Tuple[str, str]] = None
        self._req_id = 0

    # ------------------------------------------------------------------ auth

    def _get_auth(self) -> Tuple[str, str]:
        if self._user:
            return self._user, self._password

        # Cookie auth
        cookie_path = self._find_cookie()
        if cookie_path and os.path.isfile(cookie_path):
            with open(cookie_path, "r") as f:
                parts = f.read().strip().split(":")
            if len(parts) == 2:
                return parts[0], parts[1]

        if self._cookie_cache:
            return self._cookie_cache
        raise RuntimeError(
            "No RPC credentials.  Set --rpc-user/--rpc-pass or --cookie-dir."
        )

    def _find_cookie(self) -> Optional[str]:
        """Locate the .cookie file."""
        if self._cookie_dir:
            return os.path.join(self._cookie_dir, ".cookie")

        # Try common default locations
        candidates = []
        if sys.platform == "win32":
            appdata = os.environ.get("APPDATA", "")
            if appdata:
                candidates.append(os.path.join(appdata, "Meowcoin", ".cookie"))
                candidates.append(os.path.join(appdata, "Meowcoin_test", ".cookie"))
        else:
            home = os.path.expanduser("~")
            candidates.append(os.path.join(home, ".meowcoin", ".cookie"))

        for p in candidates:
            if os.path.isfile(p):
                return p
        return None

    # ------------------------------------------------------------------ call

    def call(self, method: str, params: Any = None) -> Any:
        """Make a JSON-RPC call and return the result."""
        if params is None:
            params = []
        self._req_id += 1
        payload = {
            "jsonrpc": "1.0",
            "id": self._req_id,
            "method": method,
            "params": params,
        }
        user, pw = self._get_auth()
        resp = requests.post(
            self.url,
            json=payload,
            auth=(user, pw),
            timeout=30,
        )
        data = resp.json()
        if data.get("error"):
            raise RuntimeError(f"RPC error ({method}): {data['error']}")
        return data.get("result")

    def getblocktemplate(self) -> dict:
        return self.call("getblocktemplate", [{"rules": ["segwit"]}])

    def submitblock(self, hex_data: str) -> Any:
        return self.call("submitblock", [hex_data])

# =============================================================================
# Price fetcher — MEWC/USDT from NonKYC.io
# =============================================================================

class MewcPriceFetcher:
    """
    Fetches MEWC/USDT price from the NonKYC.io public ticker endpoint.
    No API key required.  Caches the result for `cache_ttl` seconds.
    """

    TICKER_HOST = "api.nonkyc.io"
    TICKER_PATH = "/api/v2/ticker/MEWC_USDT"
    DEFAULT_CACHE_TTL = 60  # seconds

    def __init__(self, cache_ttl: int = DEFAULT_CACHE_TTL):
        self._cache_ttl = cache_ttl
        self._cached_price: Optional[float] = None
        self._cached_at: float = 0.0

    def get_price(self) -> Optional[float]:
        """Return the last MEWC/USDT price, or None on failure."""
        now = time.time()
        if self._cached_price is not None and (now - self._cached_at) < self._cache_ttl:
            return self._cached_price

        try:
            ctx = ssl.create_default_context()
            conn = http.client.HTTPSConnection(
                self.TICKER_HOST, 443, timeout=10, context=ctx,
            )
            conn.request("GET", self.TICKER_PATH,
                         headers={"Content-Type": "application/json"})
            resp = conn.getresponse()
            data = json.loads(resp.read().decode())
            conn.close()

            price = float(data.get("last_price", 0))
            if price > 0:
                self._cached_price = price
                self._cached_at = now
                log.debug("MEWC/USDT price: $%.6f", price)
                return price
        except Exception as e:
            log.warning("Failed to fetch MEWC price: %s", e)

        return self._cached_price  # return stale cache on failure

    # --- USD → CAD conversion rate ---
    FXRATE_HOST = "open.er-api.com"
    FXRATE_PATH = "/v6/latest/USD"
    FXRATE_CACHE_TTL = 3600  # 1 hour — rates don't move fast

    def __init_cad_cache(self):
        if not hasattr(self, '_cad_rate'):
            self._cad_rate: Optional[float] = None
            self._cad_cached_at: float = 0.0

    def get_usd_to_cad(self) -> Optional[float]:
        """Return the current USD→CAD exchange rate, or None on failure."""
        self.__init_cad_cache()
        now = time.time()
        if self._cad_rate is not None and (now - self._cad_cached_at) < self.FXRATE_CACHE_TTL:
            return self._cad_rate

        try:
            ctx = ssl.create_default_context()
            conn = http.client.HTTPSConnection(
                self.FXRATE_HOST, 443, timeout=10, context=ctx,
            )
            conn.request("GET", self.FXRATE_PATH)
            resp = conn.getresponse()
            data = json.loads(resp.read().decode())
            conn.close()

            rate = float(data.get("rates", {}).get("CAD", 0))
            if rate > 0:
                self._cad_rate = rate
                self._cad_cached_at = now
                log.debug("USD→CAD rate: %.4f", rate)
                return rate
        except Exception as e:
            log.warning("Failed to fetch USD→CAD rate: %s", e)

        return self._cad_rate  # return stale cache on failure


# =============================================================================
# Block logger — append to Excel on each block find
# =============================================================================

BLOCK_LOG_FILE = "block_finds.xlsx"
BLOCK_LOG_HEADERS = [
    "Date/Time (UTC)",
    "Height",
    "Block Reward (MEWC)",
    "Fees (MEWC)",
    "Total (MEWC)",
    "MEWC/USDT Price",
    "Block Value (USD)",
    "USD→CAD Rate",
    "Block Value (CAD)",
    "Coinbase TxID",
    "Worker",
    "Nonce",
    "Cumulative Blocks",
    "Cumulative MEWC",
    "Cumulative USD",
    "Cumulative CAD",
]


class BlockLogger:
    """
    Appends a row to an Excel workbook each time a block is found.
    Creates the file with headers if it does not already exist.
    """

    def __init__(self, filepath: str = BLOCK_LOG_FILE):
        self.filepath = filepath

    def log_block(
        self,
        height: int,
        reward_sat: int,
        fee_sat: int,
        price_usd: Optional[float],
        cad_rate: Optional[float],
        txid_hex: str,
        worker: str,
        nonce_hex: str,
    ):
        """Append a block-find row to the Excel file."""
        if openpyxl is None:
            log.warning("openpyxl not installed — cannot log block to Excel")
            return

        reward_mewc = reward_sat / COIN
        fee_mewc = fee_sat / COIN
        total_mewc = reward_mewc + fee_mewc
        block_usd = total_mewc * price_usd if price_usd else None
        block_cad = block_usd * cad_rate if (block_usd and cad_rate) else None
        now_utc = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

        try:
            # Load existing workbook or create new
            if os.path.exists(self.filepath):
                try:
                    wb = openpyxl.load_workbook(self.filepath)
                    ws = wb.active
                except Exception as load_err:
                    log.warning(
                        "Could not load existing %s (%s) — creating new file",
                        self.filepath, load_err,
                    )
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Block Finds"
                    ws.append(BLOCK_LOG_HEADERS)
                    from openpyxl.styles import Font, numbers
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Block Finds"
                ws.append(BLOCK_LOG_HEADERS)
                from openpyxl.styles import Font, numbers
                for cell in ws[1]:
                    cell.font = Font(bold=True)

            # Calculate cumulative totals from existing rows
            cum_blocks = ws.max_row  # header is row 1, so data rows = max_row - 1, new = max_row
            cum_mewc = total_mewc
            cum_usd = block_usd or 0.0
            cum_cad = block_cad or 0.0
            for row in ws.iter_rows(min_row=2, max_col=16, values_only=True):
                if row[4] is not None:  # Total (MEWC) column
                    cum_mewc += float(row[4])
                if row[6] is not None:  # Block Value (USD) column
                    cum_usd += float(row[6])
                if row[8] is not None:  # Block Value (CAD) column
                    try:
                        cum_cad += float(row[8])
                    except (TypeError, ValueError):
                        pass

            ws.append([
                now_utc,
                height,
                reward_mewc,
                fee_mewc,
                total_mewc,
                price_usd,
                round(block_usd, 4) if block_usd else None,
                round(cad_rate, 4) if cad_rate else None,
                round(block_cad, 4) if block_cad else None,
                txid_hex,
                worker,
                nonce_hex,
                cum_blocks,
                round(cum_mewc, 8),
                round(cum_usd, 4),
                round(cum_cad, 4),
            ])

            # Format currency columns
            row_num = ws.max_row
            ws.cell(row=row_num, column=3).number_format = '#,##0.00000000'   # Reward
            ws.cell(row=row_num, column=4).number_format = '#,##0.00000000'   # Fees
            ws.cell(row=row_num, column=5).number_format = '#,##0.00000000'   # Total
            ws.cell(row=row_num, column=6).number_format = '$#,##0.00000000'  # USD Price
            ws.cell(row=row_num, column=7).number_format = '$#,##0.0000'      # USD Value
            ws.cell(row=row_num, column=8).number_format = '#,##0.0000'       # CAD Rate
            ws.cell(row=row_num, column=9).number_format = 'C$#,##0.0000'     # CAD Value
            ws.cell(row=row_num, column=14).number_format = '#,##0.00000000'  # Cum MEWC
            ws.cell(row=row_num, column=15).number_format = '$#,##0.0000'     # Cum USD
            ws.cell(row=row_num, column=16).number_format = 'C$#,##0.0000'    # Cum CAD

            wb.save(self.filepath)
            log.info(
                "Block logged to %s: height=%d, reward=%.2f MEWC, "
                "price=$%.6f, value=$%.4f / C$%.4f",
                self.filepath, height, total_mewc,
                price_usd or 0, block_usd or 0, block_cad or 0,
            )
        except Exception as e:
            log.error("Failed to write block to Excel: %s", e)


# =============================================================================
# Discord webhook — notify on block found
# =============================================================================

class DiscordWebhook:
    """
    Sends a rich embed to a Discord channel via webhook URL whenever a block
    is found.  Silently no-ops if no webhook URL is configured.
    """

    def __init__(self, webhook_url: Optional[str] = None):
        self.webhook_url = webhook_url

    def notify_block_found(
        self,
        height: int,
        reward_mewc: float,
        fee_mewc: float,
        price_usd: Optional[float],
        cad_rate: Optional[float],
        txid_hex: str,
        worker: str,
        nonce_hex: str,
        accepted: bool,
    ):
        """Fire-and-forget Discord notification for a block event."""
        if not self.webhook_url:
            return

        try:
            total_mewc = reward_mewc + fee_mewc
            block_usd = total_mewc * price_usd if price_usd else None
            block_cad = block_usd * cad_rate if (block_usd and cad_rate) else None

            if accepted:
                color = 0x00FF00  # green
                title = f"\u26cf\ufe0f  Block Found — Height {height:,}"
            else:
                color = 0xFF0000  # red
                title = f"\u274c  Block Rejected — Height {height:,}"

            fields = [
                {"name": "Reward", "value": f"{reward_mewc:,.2f} MEWC", "inline": True},
                {"name": "Fees", "value": f"{fee_mewc:,.8f} MEWC", "inline": True},
                {"name": "Total", "value": f"{total_mewc:,.2f} MEWC", "inline": True},
            ]

            if price_usd:
                fields.append({"name": "MEWC Price", "value": f"${price_usd:,.8f}", "inline": True})
            if block_usd is not None:
                fields.append({"name": "Value (USD)", "value": f"${block_usd:,.4f}", "inline": True})
            if block_cad is not None:
                fields.append({"name": "Value (CAD)", "value": f"C${block_cad:,.4f}", "inline": True})

            fields.append({"name": "Worker", "value": worker, "inline": True})
            fields.append({"name": "Nonce", "value": f"`{nonce_hex}`", "inline": True})
            fields.append({"name": "Coinbase TxID", "value": f"`{txid_hex[:16]}...`", "inline": False})

            embed = {
                "title": title,
                "color": color,
                "fields": fields,
                "footer": {"text": "Meowcoin Solo Mining Proxy"},
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }

            payload = {
                "username": "MEWC Miner",
                "embeds": [embed],
            }

            resp = requests.post(
                self.webhook_url,
                json=payload,
                timeout=10,
            )
            if resp.status_code in (200, 204):
                log.info("Discord notification sent for block %d", height)
            else:
                log.warning("Discord webhook returned %d: %s", resp.status_code, resp.text[:200])

        except Exception as e:
            log.error("Discord webhook failed: %s", e)


# =============================================================================
# Job — represents a single mining work unit
# =============================================================================

class Job:
    """A mining job derived from a getblocktemplate response."""

    def __init__(
        self,
        job_id: str,
        version: int,
        prev_hash_internal: bytes,   # 32 bytes, internal order
        merkle_root_internal: bytes, # 32 bytes, internal order
        ntime: int,
        nbits: int,
        height: int,
        target_hex: str,             # 64-char big-endian hex from GBT
        header_hash_hex: str,        # 64-char big-endian hex
        seed_hash_hex: str,          # 64-char hex
        coinbase_raw: bytes,         # serialised coinbase tx
        tx_hex_list: List[str],      # hex-encoded non-coinbase transactions
    ):
        self.job_id = job_id
        self.version = version
        self.prev_hash_internal = prev_hash_internal
        self.merkle_root_internal = merkle_root_internal
        self.ntime = ntime
        self.nbits = nbits
        self.height = height
        self.target_hex = target_hex
        self.header_hash_hex = header_hash_hex
        self.seed_hash_hex = seed_hash_hex
        self.coinbase_raw = coinbase_raw
        self.tx_hex_list = tx_hex_list
        self.created = time.time()

    def assemble_block_hex(self, nonce64: int, mix_hash_internal: bytes) -> str:
        """
        Assemble the full serialised block as a hex string for submitblock.

        Block format (MeowPoW):
            nVersion         4B LE
            hashPrevBlock   32B internal
            hashMerkleRoot  32B internal
            nTime            4B LE
            nBits            4B LE
            nHeight          4B LE
            nNonce64         8B LE
            mix_hash        32B internal
            tx_count        varint
            transactions
        """
        parts: list[bytes] = []

        # Header
        parts.append(int32_le(self.version))
        parts.append(self.prev_hash_internal)
        parts.append(self.merkle_root_internal)
        parts.append(uint32_le(self.ntime))
        parts.append(uint32_le(self.nbits))
        parts.append(uint32_le(self.height))
        parts.append(uint64_le(nonce64))
        parts.append(mix_hash_internal)

        # Transactions
        all_tx_hex = [bytes_to_hex(self.coinbase_raw)] + self.tx_hex_list
        parts.append(compact_size(len(all_tx_hex)))
        for tx_hex in all_tx_hex:
            parts.append(hex_to_bytes(tx_hex))

        return bytes_to_hex(b"".join(parts))

# =============================================================================
# Job manager — GBT polling and job creation
# =============================================================================

class JobManager:
    """Manages block templates from the node and creates mining jobs."""

    def __init__(self, rpc: NodeRPC, mining_address: str,
                 price_fetcher: Optional[MewcPriceFetcher] = None,
                 block_logger: Optional[BlockLogger] = None,
                 discord_webhook: Optional[DiscordWebhook] = None):
        self.rpc = rpc
        self.mining_address = mining_address
        self.miner_script = address_to_scriptpubkey(mining_address)
        self.community_script = address_to_scriptpubkey(COMMUNITY_FUND_ADDRESS)
        self.jobs: Dict[str, Job] = {}
        self.current_job: Optional[Job] = None
        self._last_prev_hash: Optional[str] = None
        self._last_txs_updated: Optional[str] = None
        self._job_counter = 0
        self._last_job_time: float = 0.0
        self._max_job_age: float = 30.0  # Force new job every 30s for ntime refresh
        self._last_reconnect: float = 0.0
        self._reconnect_cooldown: float = 15.0  # Don't try reconnect more than once per 15s
        self.price_fetcher = price_fetcher or MewcPriceFetcher()
        self.block_logger = block_logger or BlockLogger()
        self.discord_webhook = discord_webhook or DiscordWebhook()

    def _try_reconnect_peers(self):
        """When GBT fails with 'not connected', try to add known peers."""
        now = time.time()
        if now - self._last_reconnect < self._reconnect_cooldown:
            return  # Don't spam reconnect attempts
        self._last_reconnect = now
        log.warning("Node has no peers — attempting to reconnect...")
        try:
            addrs = self.rpc.call("getnodeaddresses", [20])
            added = 0
            for a in addrs:
                if a.get("network") != "ipv4":
                    continue
                target = f"{a['address']}:{a['port']}"
                try:
                    self.rpc.call("addnode", [target, "onetry"])
                    added += 1
                except Exception:
                    pass
                if added >= 8:
                    break
            log.info("Sent addnode to %d peers", added)
        except Exception as e:
            log.error("Reconnect attempt failed: %s", e)

    def _next_job_id(self) -> str:
        self._job_counter += 1
        return f"{self._job_counter:08x}"

    def poll(self) -> Optional[Job]:
        """
        Poll getblocktemplate.  Returns a new Job if work changed, or if
        the current job is older than _max_job_age seconds (ntime refresh).
        """
        try:
            tpl = self.rpc.getblocktemplate()
        except Exception as e:
            err_str = str(e)
            if "not connected" in err_str.lower():
                self._try_reconnect_peers()
            else:
                log.error("GBT failed: %s", e)
            return None

        prev_hash_rpc = tpl["previousblockhash"]
        longpollid = tpl.get("longpollid", "")

        now = time.time()
        age = now - self._last_job_time
        force_refresh = age >= self._max_job_age

        # Detect new work (new tip, tx update, or stale job timeout)
        if not force_refresh and prev_hash_rpc == self._last_prev_hash and longpollid == self._last_txs_updated:
            return None

        if force_refresh and prev_hash_rpc == self._last_prev_hash:
            log.debug("Forcing job refresh after %.0fs (ntime update)", age)

        self._last_prev_hash = prev_hash_rpc
        self._last_txs_updated = longpollid
        self._last_job_time = now

        return self._create_job(tpl)

    def _create_job(self, tpl: dict) -> Job:
        # GBT version is missing top bits due to SetChainId() stripping them.
        # Meowcoin (Namecoin-derived) needs bit 28 (auxpow-aware) + bit 29
        # (BIP9 version bits) = 0x30000000 to match network blocks.
        version   = tpl["version"] | 0x30000000
        height    = tpl["height"]
        nbits     = int(tpl["bits"], 16)
        ntime     = tpl["curtime"]
        target_hex = tpl["target"]

        # Previous block hash: GBT gives big-endian hex → internal = reversed
        prev_hash_internal = hex_to_bytes(tpl["previousblockhash"])[::-1]

        # ---- Build coinbase ----
        subsidy = get_block_subsidy(height)
        community_value = (subsidy * COMMUNITY_FUND_PCT) // 100
        coinbase_value = tpl["coinbasevalue"]  # miner's share (subsidy - fund + fees)

        witness_commitment_hex = tpl.get("default_witness_commitment")
        witness_commitment = (
            hex_to_bytes(witness_commitment_hex) if witness_commitment_hex else None
        )

        coinbase_raw = build_coinbase_tx(
            height=height,
            miner_value=coinbase_value,
            miner_script=self.miner_script,
            community_value=community_value,
            community_script=self.community_script,
            witness_commitment=witness_commitment,
        )

        # ---- Compute merkle root ----
        # txid of each transaction (SHA256d of raw tx, in internal order)
        # For segwit coinbase, the txid is the hash WITHOUT witness data.
        coinbase_txid = self._txid_from_raw(coinbase_raw)
        tx_hashes = [coinbase_txid]
        tx_hex_list = []
        for tx_obj in tpl.get("transactions", []):
            tx_hex = tx_obj["data"]
            tx_hex_list.append(tx_hex)
            # Use the provided txid (it's in big-endian/display order)
            txid_display = tx_obj["txid"]
            tx_hashes.append(hex_to_bytes(txid_display)[::-1])  # → internal

        mr = merkle_root(tx_hashes)

        # ---- Compute MeowPoW header hash ----
        # SHA256d( nVersion[4] || hashPrevBlock[32] || hashMerkleRoot[32]
        #          || nTime[4] || nBits[4] || nHeight[4] )   = 80 bytes
        header_input = (
            int32_le(version)
            + prev_hash_internal
            + mr
            + uint32_le(ntime)
            + uint32_le(nbits)
            + uint32_le(height)
        )
        assert len(header_input) == 80, f"Header input is {len(header_input)} bytes, expected 80"

        header_hash_raw = sha256d(header_input)  # 32 bytes, internal order
        # For stratum and ProgPow: display as big-endian hex (reversed)
        header_hash_hex = header_hash_raw[::-1].hex()

        # ---- Epoch and seed hash ----
        epoch = height // EPOCH_LENGTH
        seed_hash = compute_seed_hash(epoch)
        seed_hash_hex = seed_hash.hex()  # raw bytes as hex (ethash convention)

        job_id = self._next_job_id()

        job = Job(
            job_id=job_id,
            version=version,
            prev_hash_internal=prev_hash_internal,
            merkle_root_internal=mr,
            ntime=ntime,
            nbits=nbits,
            height=height,
            target_hex=target_hex,
            header_hash_hex=header_hash_hex,
            seed_hash_hex=seed_hash_hex,
            coinbase_raw=coinbase_raw,
            tx_hex_list=tx_hex_list,
        )

        self.jobs[job_id] = job
        self.current_job = job

        # Prune old jobs (keep last 10)
        if len(self.jobs) > 10:
            old_ids = sorted(self.jobs.keys())[:-10]
            for oid in old_ids:
                del self.jobs[oid]

        log.info(
            "New job %s  height=%d  epoch=%d  txs=%d  target=%s...",
            job_id, height, epoch, len(tx_hex_list), target_hex[:16],
        )

        return job

    @staticmethod
    def _txid_from_raw(raw_tx: bytes) -> bytes:
        """
        Compute the txid from a serialised transaction.
        For segwit transactions, strip the witness data first.
        """
        # Check for segwit marker
        if len(raw_tx) > 6 and raw_tx[4] == 0x00 and raw_tx[5] == 0x01:
            # Segwit tx — strip marker, flag, and witness
            # nVersion (4) + marker (1) + flag (1) + rest
            n_version = raw_tx[:4]
            rest = raw_tx[6:]  # after marker+flag

            # Read inputs
            pos = 0
            n_vin, sz = _read_varint(rest, pos)
            pos = sz
            for _ in range(n_vin):
                pos += 32 + 4  # prev_hash + prev_index
                script_len, sz = _read_varint(rest, pos)
                pos = sz + script_len + 4  # scriptSig + nSequence

            # Read outputs
            n_vout, sz = _read_varint(rest, pos)
            pos = sz
            for _ in range(n_vout):
                pos += 8  # nValue
                script_len, sz = _read_varint(rest, pos)
                pos = sz + script_len

            # Everything from here to (end - 4) is witness data
            # nLockTime is the last 4 bytes
            n_locktime = raw_tx[-4:]
            stripped = n_version + rest[:pos] + n_locktime
            return sha256d(stripped)
        else:
            return sha256d(raw_tx)

    def submit_solution(self, job_id: str, nonce_hex: str, mix_hash_hex: str,
                         worker: str = "unknown") -> str:
        """
        Assemble the block from a mining solution and submit it.
        Returns the submitblock result (None = accepted, string = error).
        """
        t0 = time.time()
        job = self.jobs.get(job_id)
        if not job:
            return "job-not-found"

        nonce64 = int(nonce_hex, 16)
        # mix_hash from miner is big-endian hex → internal = reversed
        mix_hash_internal = hex_to_bytes(mix_hash_hex)[::-1]

        block_hex = job.assemble_block_hex(nonce64, mix_hash_internal)
        t1 = time.time()

        log.info(
            "Submitting block  job=%s  height=%d  nonce=%s  (assembled in %.3fs)",
            job_id, job.height, nonce_hex, t1 - t0,
        )

        try:
            result = self.rpc.submitblock(block_hex)
        except Exception as e:
            log.error("submitblock RPC failed: %s", e)
            return str(e)

        t2 = time.time()
        # submitblock returns null on success; some node versions return ""
        if result is None or result == "":
            # --- Block accepted! Fetch price and log to Excel ---
            try:
                price = self.price_fetcher.get_price()
                cad_rate = self.price_fetcher.get_usd_to_cad()
                subsidy = get_block_subsidy(job.height)
                community_share = (subsidy * COMMUNITY_FUND_PCT) // 100
                miner_reward = subsidy - community_share
                fee_sat = 0  # future: extract from template

                # Compute the coinbase transaction ID
                coinbase_txid = self._txid_from_raw(job.coinbase_raw)
                txid_hex = coinbase_txid[::-1].hex()  # display order (big-endian)

                log.info(
                    "*** BLOCK ACCEPTED ***  height=%d  reward=%.2f MEWC  price=$%.6f  "
                    "CAD rate=%.4f  txid=%s  (submitblock took %.3fs, total %.3fs)",
                    job.height, miner_reward / COIN,
                    price or 0, cad_rate or 0, txid_hex, t2 - t1, t2 - t0,
                )

                self.block_logger.log_block(
                    height=job.height,
                    reward_sat=miner_reward,
                    fee_sat=fee_sat,
                    price_usd=price,
                    cad_rate=cad_rate,
                    txid_hex=txid_hex,
                    worker=worker,
                    nonce_hex=nonce_hex,
                )

                self.discord_webhook.notify_block_found(
                    height=job.height,
                    reward_mewc=miner_reward / COIN,
                    fee_mewc=fee_sat / COIN,
                    price_usd=price,
                    cad_rate=cad_rate,
                    txid_hex=txid_hex,
                    worker=worker,
                    nonce_hex=nonce_hex,
                    accepted=True,
                )
            except Exception as e:
                log.error(
                    "Block accepted at height %d but post-accept processing "
                    "failed: %s\n%s", job.height, e, traceback.format_exc(),
                )
        else:
            log.warning("Block rejected: %s  (submitblock took %.3fs)", result, t2 - t1)

            self.discord_webhook.notify_block_found(
                height=job.height,
                reward_mewc=0,
                fee_mewc=0,
                price_usd=None,
                cad_rate=None,
                txid_hex="",
                worker=worker,
                nonce_hex=nonce_hex,
                accepted=False,
            )

        return "accepted" if (result is None or result == "") else result


def _read_varint(data: bytes, pos: int) -> Tuple[int, int]:
    """Read a CompactSize varint from data at pos.  Returns (value, new_pos)."""
    b = data[pos]
    if b < 0xFD:
        return b, pos + 1
    elif b == 0xFD:
        return struct.unpack_from("<H", data, pos + 1)[0], pos + 3
    elif b == 0xFE:
        return struct.unpack_from("<I", data, pos + 1)[0], pos + 5
    else:
        return struct.unpack_from("<Q", data, pos + 1)[0], pos + 9

# =============================================================================
# Stratum protocol — miner session
# =============================================================================

class MinerSession:
    """Handles a single TCP miner connection."""

    def __init__(
        self,
        reader: asyncio.StreamReader,
        writer: asyncio.StreamWriter,
        job_manager: JobManager,
        on_disconnect,
    ):
        self.reader = reader
        self.writer = writer
        self.job_manager = job_manager
        self.on_disconnect = on_disconnect
        self.peer = writer.get_extra_info("peername")
        self.subscription_id = uuid.uuid4().hex[:16]
        self.authorized = False
        self.worker_name = "unknown"
        self._closed = False

    async def run(self):
        """Main read loop for this miner connection."""
        log.info("Miner connected: %s", self.peer)
        try:
            while not self._closed:
                line = await self.reader.readline()
                if not line:
                    break
                line = line.strip()
                if not line:
                    continue
                try:
                    msg = json.loads(line.decode("utf-8", errors="replace"))
                except json.JSONDecodeError:
                    log.warning("Invalid JSON from %s: %s", self.peer, line[:200])
                    continue
                await self._handle_message(msg)
        except (asyncio.IncompleteReadError, ConnectionResetError, BrokenPipeError):
            pass
        except Exception:
            log.error("Session error (%s): %s", self.peer, traceback.format_exc())
        finally:
            await self.close()
            log.info("Miner disconnected: %s", self.peer)
            self.on_disconnect(self)

    async def _handle_message(self, msg: dict):
        method = msg.get("method", "")
        msg_id = msg.get("id")
        params = msg.get("params", [])

        if method == "mining.subscribe":
            await self._handle_subscribe(msg_id, params)
        elif method == "mining.authorize":
            await self._handle_authorize(msg_id, params)
        elif method == "mining.submit":
            await self._handle_submit(msg_id, params)
        elif method == "mining.extranonce.subscribe":
            await self._send_result(msg_id, True)
        else:
            log.debug("Unknown method from %s: %s", self.peer, method)
            await self._send_error(msg_id, 20, f"Unknown method: {method}")

    async def _handle_subscribe(self, msg_id, params):
        miner_agent = params[0] if params else "unknown"
        log.info("Subscribe from %s: agent=%s", self.peer, miner_agent)

        # KawPow subscribe response: [session_id, extranonce1]
        # Reference: RavenCommunity/kawpow-stratum-pool stratum.js
        result = [
            None,                          # session id (null)
            self.subscription_id[:8],      # extranonce1
        ]
        await self._send_result(msg_id, result)

        # Send initial target
        if self.job_manager.current_job:
            await self.send_set_target(self.job_manager.current_job.target_hex)
            await self.send_notify(self.job_manager.current_job, clean=True)

    async def _handle_authorize(self, msg_id, params):
        self.worker_name = params[0] if params else "unknown"
        self.authorized = True
        log.info("Authorized: %s (%s)", self.worker_name, self.peer)
        await self._send_result(msg_id, True)

        # Send current job after authorization
        if self.job_manager.current_job:
            await self.send_set_target(self.job_manager.current_job.target_hex)
            await self.send_notify(self.job_manager.current_job, clean=True)

    async def _handle_submit(self, msg_id, params):
        """
        Handle mining.submit from miner.
        Expected params: [worker, job_id, nonce_hex, header_hash, mix_hash]
        Some miners may omit header_hash: [worker, job_id, nonce_hex, mix_hash]
        """
        if not self.authorized:
            await self._send_error(msg_id, 24, "Not authorized")
            return

        if len(params) < 4:
            await self._send_error(msg_id, 21, "Not enough parameters")
            return

        worker = params[0]
        job_id = params[1]
        # Strip 0x prefix if present (kawpow miners send 0x-prefixed values)
        nonce_hex = params[2]
        if nonce_hex.startswith("0x") or nonce_hex.startswith("0X"):
            nonce_hex = nonce_hex[2:]

        # Detect format: 5 params = [worker, job_id, nonce, header_hash, mix_hash]
        #                4 params = [worker, job_id, nonce, mix_hash]
        if len(params) >= 5:
            mix_hash_hex = params[4]
        else:
            mix_hash_hex = params[3]

        # Strip 0x prefix from mix_hash if present
        if mix_hash_hex.startswith("0x") or mix_hash_hex.startswith("0X"):
            mix_hash_hex = mix_hash_hex[2:]

        log.info(
            "Share from %s: job=%s nonce=%s mix=%s...",
            worker, job_id, nonce_hex, mix_hash_hex[:16],
        )

        # Run submitblock in a thread so we don't block the event loop
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None,
            self.job_manager.submit_solution,
            job_id, nonce_hex, mix_hash_hex, worker,
        )

        if result == "accepted":
            await self._send_result(msg_id, True)
        elif result == "job-not-found":
            await self._send_error(msg_id, 21, "Job not found (stale)")
        else:
            await self._send_error(msg_id, 20, f"Rejected: {result}")

    async def send_set_target(self, target_hex: str):
        """Send mining.set_target to the miner."""
        await self._send_notification("mining.set_target", [target_hex])

    async def send_notify(self, job: Job, clean: bool = True):
        """
        Send mining.notify with a new job.
        Params: [job_id, header_hash, seed_hash, target, clean_jobs, height, bits]
        Reference: RavenCommunity/kawpow-stratum-pool blockTemplate.js getJobParams()
        - height MUST be an integer (not hex string) — SRBMiner parses it as int
        - bits is a hex string from getblocktemplate rpcData
        """
        params = [
            job.job_id,
            job.header_hash_hex,
            job.seed_hash_hex,
            job.target_hex,
            clean,
            job.height,                    # integer, NOT hex string
            format(job.nbits, "08x"),      # bits as hex string
        ]
        await self._send_notification("mining.notify", params)

    # ------------------------------------------------------------------ I/O

    async def _send_result(self, msg_id, result):
        await self._send_json({"id": msg_id, "result": result, "error": None})

    async def _send_error(self, msg_id, code: int, message: str):
        await self._send_json({"id": msg_id, "result": None, "error": [code, message, None]})

    async def _send_notification(self, method: str, params):
        await self._send_json({"id": None, "method": method, "params": params})

    async def _send_json(self, obj):
        if self._closed:
            return
        try:
            line = json.dumps(obj) + "\n"
            self.writer.write(line.encode("utf-8"))
            await self.writer.drain()
        except (ConnectionResetError, BrokenPipeError, OSError):
            self._closed = True

    async def close(self):
        if not self._closed:
            self._closed = True
            try:
                self.writer.close()
                await self.writer.wait_closed()
            except Exception:
                pass

# =============================================================================
# Stratum server — main loop
# =============================================================================

class StratumServer:
    """
    Async Stratum server.  Accepts miner connections and polls
    the node for new work.
    """

    def __init__(
        self,
        job_manager: JobManager,
        host: str = "0.0.0.0",
        port: int = 3333,
        poll_interval: float = 1.0,
    ):
        self.job_manager = job_manager
        self.host = host
        self.port = port
        self.poll_interval = poll_interval
        self.sessions: List[MinerSession] = []
        self._server: Optional[asyncio.AbstractServer] = None

    async def start(self):
        self._server = await asyncio.start_server(
            self._handle_client, self.host, self.port,
        )
        addrs = ", ".join(str(s.getsockname()) for s in self._server.sockets)
        log.info("Stratum server listening on %s", addrs)
        log.info("Mining address: %s", self.job_manager.mining_address)

        # Start the GBT polling loop as a background task
        asyncio.create_task(self._poll_loop())

        async with self._server:
            await self._server.serve_forever()

    async def _handle_client(self, reader: asyncio.StreamReader, writer: asyncio.StreamWriter):
        session = MinerSession(reader, writer, self.job_manager, self._on_disconnect)
        self.sessions.append(session)
        await session.run()

    def _on_disconnect(self, session: MinerSession):
        if session in self.sessions:
            self.sessions.remove(session)

    async def _poll_loop(self):
        """Periodically poll GBT and broadcast new jobs to all miners."""
        while True:
            try:
                t0 = time.time()
                job = await asyncio.get_event_loop().run_in_executor(
                    None, self.job_manager.poll,
                )
                if job:
                    t1 = time.time()
                    await self._broadcast(job)
                    t2 = time.time()
                    log.debug(
                        "Poll+broadcast: GBT=%.3fs broadcast=%.3fs total=%.3fs",
                        t1 - t0, t2 - t1, t2 - t0,
                    )
                    # After a new block, immediately re-poll instead of sleeping
                    # to catch rapid successive blocks
                    continue
            except Exception:
                log.error("Poll error: %s", traceback.format_exc())

            await asyncio.sleep(self.poll_interval)

    async def _broadcast(self, job: Job):
        """Send a new job to all connected miners."""
        for session in list(self.sessions):
            try:
                await session.send_set_target(job.target_hex)
                await session.send_notify(job, clean=True)
            except Exception:
                pass

# =============================================================================
# Main
# =============================================================================

def parse_args():
    p = argparse.ArgumentParser(
        description="Meowcoin MeowPoW Solo Mining Stratum Proxy",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--address", default="", help="Meowcoin payout address")
    p.add_argument("--rpc-host", default="127.0.0.1", help="Node RPC host (default: 127.0.0.1)")
    p.add_argument("--rpc-port", type=int, default=8332, help="Node RPC port (default: 8332)")
    p.add_argument("--rpc-user", default="", help="Node RPC username")
    p.add_argument("--rpc-pass", default="", help="Node RPC password")
    p.add_argument("--cookie-dir", default="", help="Data directory containing .cookie file")
    p.add_argument("--stratum-host", default="0.0.0.0", help="Stratum listen host (default: 0.0.0.0)")
    p.add_argument("--stratum-port", type=int, default=3333, help="Stratum listen port (default: 3333)")
    p.add_argument("--poll-interval", type=float, default=1.0, help="GBT poll interval in seconds (default: 1.0)")
    p.add_argument("--log-level", default="INFO", choices=["DEBUG", "INFO", "WARNING", "ERROR"])
    p.add_argument("--block-log", default="block_finds.xlsx",
                   help="Excel file for block find log (default: block_finds.xlsx)")
    p.add_argument("--discord-webhook", default="",
                   help="Discord webhook URL for block-found notifications")
    return p.parse_args()


def _is_frozen():
    """Return True if running as a PyInstaller bundle."""
    return getattr(sys, 'frozen', False)


def _pause_and_exit(code: int = 1):
    """Pause so double-clicked exe users can read the error, then exit."""
    if _is_frozen():
        print()
        input("Press Enter to exit...")
    sys.exit(code)


def _interactive_setup(args):
    """Prompt for required settings when launched without CLI arguments (e.g. double-click)."""
    print("="*60)
    print("  Meowcoin MeowPoW Solo Mining Stratum Proxy  v1.04")
    print("="*60)
    print()

    if not args.address:
        args.address = input("  Mining address: ").strip()
        if not args.address:
            print("\n  ERROR: A Meowcoin address is required.")
            _pause_and_exit(1)

    rpc_host = input(f"  Node RPC host [{args.rpc_host}]: ").strip()
    if rpc_host:
        args.rpc_host = rpc_host

    rpc_port = input(f"  Node RPC port [{args.rpc_port}]: ").strip()
    if rpc_port:
        try:
            args.rpc_port = int(rpc_port)
        except ValueError:
            print(f"  Invalid port '{rpc_port}', using default {args.rpc_port}")

    rpc_user = input(f"  RPC username (blank=cookie auth) [{args.rpc_user or ''}]: ").strip()
    if rpc_user:
        args.rpc_user = rpc_user

    rpc_pass = input(f"  RPC password (blank=cookie auth) [{args.rpc_pass or ''}]: ").strip()
    if rpc_pass:
        args.rpc_pass = rpc_pass

    stratum_port = input(f"  Stratum listen port [{args.stratum_port}]: ").strip()
    if stratum_port:
        try:
            args.stratum_port = int(stratum_port)
        except ValueError:
            print(f"  Invalid port '{stratum_port}', using default {args.stratum_port}")

    print()
    return args


def main():
    args = parse_args()

    # If no address provided (e.g. double-clicked exe), run interactive setup
    if not args.address:
        args = _interactive_setup(args)

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(asctime)s  %(name)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Validate mining address
    try:
        script = address_to_scriptpubkey(args.address)
        log.info("Mining address %s → scriptPubKey %s", args.address, script.hex())
    except Exception as e:
        log.error("Invalid mining address '%s': %s", args.address, e)
        _pause_and_exit(1)

    rpc = NodeRPC(
        host=args.rpc_host,
        port=args.rpc_port,
        user=args.rpc_user,
        password=args.rpc_pass,
        cookie_dir=args.cookie_dir,
    )

    # Verify RPC connectivity
    try:
        info = rpc.call("getmininginfo")
        log.info(
            "Connected to node — chain=%s  height=%s  difficulty=%s",
            info.get("chain", "?"),
            info.get("blocks", "?"),
            info.get("difficulty", "?"),
        )
    except Exception as e:
        log.error("Cannot connect to node RPC: %s", e)
        _pause_and_exit(1)

    price_fetcher = MewcPriceFetcher()
    block_logger = BlockLogger(filepath=args.block_log)

    # Fetch initial MEWC price and CAD rate
    price = price_fetcher.get_price()
    cad_rate = price_fetcher.get_usd_to_cad()
    if price:
        log.info("MEWC/USDT price: $%.6f", price)
    else:
        log.warning("Could not fetch MEWC price — will retry on block find")
    if cad_rate:
        log.info("USD→CAD rate: %.4f", cad_rate)
    else:
        log.warning("Could not fetch USD→CAD rate — will retry on block find")

    log.info("Block finds will be logged to: %s", os.path.abspath(args.block_log))

    discord_webhook = DiscordWebhook(webhook_url=args.discord_webhook or None)
    if args.discord_webhook:
        log.info("Discord webhook notifications enabled")

    job_mgr = JobManager(rpc, args.address,
                         price_fetcher=price_fetcher,
                         block_logger=block_logger,
                         discord_webhook=discord_webhook)
    server = StratumServer(
        job_manager=job_mgr,
        host=args.stratum_host,
        port=args.stratum_port,
        poll_interval=args.poll_interval,
    )

    log.info(
        "Starting stratum proxy on %s:%d  →  node %s:%d",
        args.stratum_host, args.stratum_port,
        args.rpc_host, args.rpc_port,
    )

    try:
        asyncio.run(server.start())
    except KeyboardInterrupt:
        log.info("Shutting down.")
    except Exception as e:
        log.error("Fatal error: %s", e)
        _pause_and_exit(1)


if __name__ == "__main__":
    main()
